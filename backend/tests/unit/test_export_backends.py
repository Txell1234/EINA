"""
Unit tests for export_backends (Excel + WeasyPrint probes)
"""
from pathlib import Path

import pytest

from services.export_backends import (
    probe_openpyxl,
    probe_pdf_renderers,
    probe_playwright,
    probe_weasyprint,
    render_pdf_from_html,
    write_case_report_excel,
)


SAMPLE_REPORT = {
    "case": {"id": 1, "name": "Cas prova", "description": "Descripció de prova"},
    "osint_data": [{"id": 10, "type": "news", "status": "completed"}],
    "ai_analyses": [{"id": 2, "type": "geopolitical", "confidence": 0.85}],
    "qualitative_analyses": [],
    "predictions": [{"id": 3, "type": "trend", "confidence": 72}],
    "investment_recommendations": [],
    "premises": [{"id": 1, "premise_text": "Premissa A", "framework_id": 1, "created_at": "2026-01-01"}],
    "bias_guidance": {"enabled": True, "premise_count": 1, "notes": ["Premissa A"]},
}


@pytest.mark.unit
class TestExportBackends:
    def test_probe_openpyxl_available(self):
        status = probe_openpyxl()
        assert status["available"] is True
        assert "version" in status

    def test_probe_weasyprint_returns_dict(self):
        status = probe_weasyprint(smoke_test=False)
        assert "available" in status
        assert isinstance(status["available"], bool)

    def test_probe_playwright_returns_dict(self):
        status = probe_playwright(smoke_test=False)
        assert "available" in status
        assert isinstance(status["available"], bool)

    def test_probe_pdf_renderers_aggregate(self):
        status = probe_pdf_renderers(smoke_test=False)
        assert "available" in status
        assert "weasyprint" in status
        assert "playwright" in status

    @pytest.mark.slow
    def test_render_pdf_playwright_if_available(self, tmp_path: Path):
        pw = probe_playwright(smoke_test=False)
        if not pw.get("available"):
            pytest.skip("Playwright no disponible")
        out = tmp_path / "probe.pdf"
        render_pdf_from_html("<html><body><h1>EINA</h1></body></html>", out)
        assert out.exists()
        assert out.stat().st_size > 500

    def test_write_case_report_excel_creates_workbook(self, tmp_path: Path):
        out = tmp_path / "report_1.xlsx"
        write_case_report_excel(SAMPLE_REPORT, out)

        assert out.exists()
        assert out.stat().st_size > 500

        from openpyxl import load_workbook

        wb = load_workbook(out, read_only=True)
        assert "Cas" in wb.sheetnames
        assert "OSINT" in wb.sheetnames
        assert "Anàlisi IA" in wb.sheetnames
        assert "Guia biaix" in wb.sheetnames
        wb.close()

    def test_write_case_report_excel_empty_sections(self, tmp_path: Path):
        data = {
            "case": {"id": 2, "name": "Buit", "description": ""},
            "osint_data": [],
            "ai_analyses": [],
            "qualitative_analyses": [],
            "predictions": [],
            "investment_recommendations": [],
            "premises": [],
            "bias_guidance": {"enabled": False, "premise_count": 0, "notes": []},
        }
        out = tmp_path / "empty.xlsx"
        write_case_report_excel(data, out)
        assert out.exists()

    @pytest.mark.slow
    def test_weasyprint_smoke_test_if_available(self):
        status = probe_weasyprint(smoke_test=True)
        if status.get("available"):
            assert status.get("error") is None
        else:
            assert "install_hint" in status

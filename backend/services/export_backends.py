"""
Shared export backends — WeasyPrint (PDF) and openpyxl (Excel).

Centralises dependency probes and install hints for production deployments.
"""
from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# Debian/Ubuntu packages required for WeasyPrint native bindings
WEASYPRINT_APT_PACKAGES = (
    "libpango-1.0-0",
    "libpangocairo-1.0-0",
    "libgdk-pixbuf-2.0-0",
    "libffi-dev",
    "shared-mime-info",
    "fonts-dejavu-core",
)

WEASYPRINT_INSTALL_HINT = (
    "Instal·la dependències natives: "
    f"apt-get install -y {' '.join(WEASYPRINT_APT_PACKAGES)} "
    "(Linux) o usa backend/Dockerfile. "
    "Windows: https://doc.courtbouillon.org/weasyprint/stable/first_steps.html"
)

PLAYWRIGHT_INSTALL_HINT = (
    "pip install playwright && python -m playwright install chromium "
    "(fallback PDF quan WeasyPrint no està disponible, p.ex. Windows local)"
)


class ExportBackendError(Exception):
    """Raised when an export backend is unavailable or fails."""

    def __init__(self, backend: str, reason: str, *, hint: str | None = None):
        self.backend = backend
        self.reason = reason
        self.hint = hint or WEASYPRINT_INSTALL_HINT
        super().__init__(f"{backend}: {reason}")


def probe_weasyprint(*, smoke_test: bool = False) -> dict[str, Any]:
    """
    Check whether WeasyPrint can be used.
    smoke_test=True runs a minimal in-memory PDF (slower, use in /health or CI).
    """
    try:
        import weasyprint  # noqa: F401
        from weasyprint import HTML

        version = getattr(weasyprint, "__version__", "unknown")
        if smoke_test:
            buf = io.BytesIO()
            HTML(string="<html><body><p>EINA</p></body></html>").write_pdf(buf)
            if len(buf.getvalue()) < 100:
                return {
                    "available": False,
                    "version": version,
                    "error": "smoke_test_failed",
                    "message": "PDF generat buit",
                    "install_hint": WEASYPRINT_INSTALL_HINT,
                }
        return {
            "available": True,
            "version": version,
            "install_hint": WEASYPRINT_INSTALL_HINT,
        }
    except ImportError as exc:
        return {
            "available": False,
            "error": "import",
            "message": str(exc),
            "install_hint": "pip install 'weasyprint>=60.0'",
        }
    except OSError as exc:
        return {
            "available": False,
            "error": "native_libs",
            "message": str(exc),
            "install_hint": WEASYPRINT_INSTALL_HINT,
        }
    except Exception as exc:
        return {
            "available": False,
            "error": "runtime",
            "message": str(exc),
            "install_hint": WEASYPRINT_INSTALL_HINT,
        }


def probe_playwright(*, smoke_test: bool = False) -> dict[str, Any]:
    """Chromium headless PDF via Playwright (works on Windows without GTK/Cairo)."""
    try:
        from playwright.sync_api import sync_playwright

        version = getattr(sync_playwright, "__version__", "unknown")
        if smoke_test:
            buf = io.BytesIO()
            with sync_playwright() as p:
                browser = p.chromium.launch()
                page = browser.new_page()
                page.set_content("<html><body><p>EINA</p></body></html>", wait_until="networkidle")
                data = page.pdf(format="A4", print_background=True)
                browser.close()
            buf.write(data)
            if len(buf.getvalue()) < 100:
                return {
                    "available": False,
                    "version": version,
                    "error": "smoke_test_failed",
                    "message": "PDF generat buit",
                    "install_hint": PLAYWRIGHT_INSTALL_HINT,
                }
        return {
            "available": True,
            "version": version,
            "install_hint": PLAYWRIGHT_INSTALL_HINT,
        }
    except ImportError as exc:
        return {
            "available": False,
            "error": "import",
            "message": str(exc),
            "install_hint": PLAYWRIGHT_INSTALL_HINT,
        }
    except Exception as exc:
        return {
            "available": False,
            "error": "runtime",
            "message": str(exc),
            "install_hint": PLAYWRIGHT_INSTALL_HINT,
        }


def probe_pdf_renderers(*, smoke_test: bool = False) -> dict[str, Any]:
    """Aggregate PDF backend availability (WeasyPrint preferred, Playwright fallback)."""
    weasy = probe_weasyprint(smoke_test=smoke_test)
    playwright = probe_playwright(smoke_test=smoke_test)
    preferred = None
    if weasy.get("available"):
        preferred = "weasyprint"
    elif playwright.get("available"):
        preferred = "playwright"
    return {
        "available": bool(preferred),
        "preferred": preferred,
        "weasyprint": weasy,
        "playwright": playwright,
    }


def _render_pdf_weasyprint(html: str, path: Path, *, base_url: str | None = None) -> None:
    from weasyprint import HTML

    base = base_url or str(path.parent.resolve())
    HTML(string=html, base_url=base).write_pdf(str(path))


def _render_pdf_playwright(html: str, path: Path) -> None:
    from playwright.sync_api import sync_playwright

    path.parent.mkdir(parents=True, exist_ok=True)
    with sync_playwright() as p:
        browser = p.chromium.launch()
        page = browser.new_page()
        page.set_content(html, wait_until="networkidle")
        page.pdf(path=str(path), format="A4", print_background=True, margin={"top": "12mm", "bottom": "12mm"})
        browser.close()


def probe_openpyxl() -> dict[str, Any]:
    try:
        import openpyxl

        return {
            "available": True,
            "version": getattr(openpyxl, "__version__", "unknown"),
        }
    except ImportError as exc:
        return {
            "available": False,
            "error": "import",
            "message": str(exc),
            "install_hint": "pip install 'openpyxl>=3.1.0'",
        }


def render_pdf_from_html(html: str, path: Path, *, base_url: str | None = None) -> None:
    """Write PDF to path using WeasyPrint or Playwright fallback."""
    weasy = probe_weasyprint(smoke_test=False)
    if weasy.get("available"):
        try:
            _render_pdf_weasyprint(html, path, base_url=base_url)
            return
        except Exception as exc:
            logger.warning("WeasyPrint PDF failed, trying Playwright: %s", exc)

    playwright = probe_playwright(smoke_test=False)
    if playwright.get("available"):
        try:
            _render_pdf_playwright(html, path)
            return
        except Exception as exc:
            raise ExportBackendError(
                "playwright",
                str(exc),
                hint=playwright.get("install_hint"),
            ) from exc

    hint = (
        f"{weasy.get('install_hint', WEASYPRINT_INSTALL_HINT)} · "
        f"Alternativa: {PLAYWRIGHT_INSTALL_HINT}"
    )
    raise ExportBackendError(
        "pdf",
        weasy.get("message") or playwright.get("message") or "Cap motor PDF disponible",
        hint=hint,
    )


def _sheet_title(name: str) -> str:
    """Excel sheet names max 31 chars, no : \\ / ? * [ ]"""
    cleaned = name.replace(":", " ").replace("/", "-")[:31]
    return cleaned or "Sheet"


def _append_table_sheet(wb, title: str, rows: list[dict], columns: list[tuple[str, str]]) -> None:
    from openpyxl.styles import Font

    ws = wb.create_sheet(_sheet_title(title))
    headers = [label for _, label in columns]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    if not rows:
        ws.append(["Sense dades"])
        return
    for row in rows:
        ws.append([row.get(key) for key, _ in columns])


def write_case_report_excel(data: dict, path: Path) -> None:
    """Build multi-sheet .xlsx case report. Raises ExportBackendError if openpyxl missing."""
    status = probe_openpyxl()
    if not status.get("available"):
        raise ExportBackendError(
            "openpyxl",
            status.get("message", "no disponible"),
            hint=status.get("install_hint"),
        )

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    case = data.get("case") or {}
    ws = wb.active
    ws.title = "Cas"
    ws["A1"] = "Informe OSINT — EINA"
    ws["A1"].font = Font(bold=True, size=14)
    for row_idx, (label, value) in enumerate(
        [
            ("ID cas", case.get("id")),
            ("Nom", case.get("name")),
            ("Descripció", case.get("description")),
        ],
        start=3,
    ):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)

    sections: list[tuple[str, str, list[tuple[str, str]]]] = [
        ("OSINT", "osint_data", [("id", "ID"), ("type", "Tipus"), ("status", "Estat")]),
        (
            "Anàlisi IA",
            "ai_analyses",
            [("id", "ID"), ("type", "Tipus"), ("confidence", "Confiança")],
        ),
        (
            "Qualitativa",
            "qualitative_analyses",
            [("id", "ID"), ("conclusions", "Conclusions"), ("confidence", "Confiança")],
        ),
        (
            "Prediccions",
            "predictions",
            [("id", "ID"), ("type", "Tipus"), ("confidence", "Confiança %")],
        ),
        (
            "Inversió",
            "investment_recommendations",
            [("id", "ID"), ("type", "Tipus"), ("confidence", "Confiança %")],
        ),
        (
            "Premisses",
            "premises",
            [
                ("id", "ID"),
                ("premise_text", "Premissa"),
                ("framework_id", "Marc"),
                ("created_at", "Creat"),
            ],
        ),
    ]
    for title, key, columns in sections:
        _append_table_sheet(wb, title, data.get(key) or [], columns)

    bias = data.get("bias_guidance") or {}
    bias_ws = wb.create_sheet("Guia biaix")
    bias_ws["A1"] = "Guia de biaix"
    bias_ws["A1"].font = Font(bold=True, size=12)
    bias_ws["A3"] = "Activada"
    bias_ws["B3"] = bias.get("enabled")
    bias_ws["A4"] = "Nombre premisses"
    bias_ws["B4"] = bias.get("premise_count")
    notes = bias.get("notes") or []
    if notes:
        bias_ws["A6"] = "Notes"
        bias_ws["A6"].font = Font(bold=True)
        for i, note in enumerate(notes, start=7):
            bias_ws.cell(row=i, column=1, value=note)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    logger.info("Excel report written to %s", path)
